import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from core.conf import translator_service


def format_partner_balance_excel(data: list, lang: str = "ru", translator=translator_service,
                                 show_firm: bool = True, newest_first: bool = False,
                                 output_dir: str = "exports", header_text: str = None) -> str:
    """
    Exports partner balance data to Excel file.
    Creates a separate sheet for each currency.

    Parameters:
        data: list of dicts with document info
        lang: 'en', 'ru', 'uz'
        translator: object with translator.get(key, lang)
        show_firm: include firm header if True
        newest_first: reverse order of operations
        output_dir: directory to save the file
        header_text: optional custom header text to display at the top

    Returns:
        str: path to the created Excel file
    """
    if not data:
        raise ValueError("No data to export")

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Determine firm name
    firm_name = data[0].get("firm", {}).get("name", "")

    # Group operations by currency
    currencies = {}
    for op in data:
        currency_name = op.get("currency", {}).get("name", "Unknown")
        currencies.setdefault(currency_name, []).append(op)

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    firm_font = Font(bold=True, size=14)
    custom_header_font = Font(bold=True, size=16, color="FFFFFF")
    custom_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for currency_name, operations in currencies.items():
        # Sort by date
        operations.sort(key=lambda x: x.get("date", 0), reverse=newest_first)

        # Create sheet for currency
        sheet_name = currency_name[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        row = 1

        # Add custom header if provided
        if header_text:
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = header_text
            cell.font = custom_header_font
            cell.fill = custom_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 2

        # Add firm header if needed
        if show_firm and firm_name:
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = f"🏢 {firm_name}"
            cell.font = firm_font
            cell.alignment = Alignment(horizontal='center')
            row += 2

        # Add currency header
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"💱 {currency_name}"
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Add column headers
        headers = [
            translator.get('document_type', lang),
            translator.get('document_code', lang),
            translator.get('exchange_rate', lang),
            translator.get('debit', lang),
            translator.get('credit', lang),
            translator.get('remainder', lang),
            translator.get('date', lang)
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row += 1

        # Add data rows
        last_op = operations[-1]
        for op in operations:
            doc_type_id = op.get("document_type", {}).get("id", "0")
            code = op.get("document_code", "—")
            curr = op.get("currency", {})
            exch = curr.get("exchange_rate", 1)
            start = op.get("start_amount", 0)
            debit = op.get("debit", 0)
            credit = op.get("credit", 0)
            date = datetime.fromtimestamp(op.get("date", 0)).strftime("%d.%m.%y %H:%M")
            remainder = start + debit - credit

            doc_type_name = translator.get(f"partner_document_type{doc_type_id}", lang)

            # Mark current remainder
            remainder_label = f"{remainder:,.2f}"
            if op is last_op:
                remainder_label += f" ({translator.get('current', lang)})"

            row_data = [
                doc_type_name,
                code,
                f"{exch:,.2f}" if exch != 1 else "1.00",
                f"{debit:,.2f}" if debit != 0 else "—",
                f"{credit:,.2f}" if credit != 0 else "—",
                remainder_label,
                date
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col in [1, 2, 7] else 'right')

                # Highlight current remainder
                if col == 6 and op is last_op:
                    cell.font = bold_font
                    cell.fill = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")

            row += 1

        # Adjust column widths
        column_widths = [25, 15, 15, 15, 15, 20, 18]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"partner_balance_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Save workbook
    wb.save(filepath)

    return filepath


def format_total_excel(data: list, lang: str = "ru", translator=translator_service,
                       output_dir: str = "exports", header_text: str = None) -> str:
    """
    Exports partner balance totals to Excel file.
    Groups by document_type, then by currency.
    Shows totals per document_type and overall total per currency.

    Parameters:
        data: list of dicts with document info
        lang: 'en', 'ru', 'uz'
        translator: object with translator.get(key, lang)
        output_dir: directory to save the file
        header_text: optional custom header text to display at the top

    Returns:
        str: path to the created Excel file
    """
    if not data:
        raise ValueError("No data to export")

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = translator.get('total', lang)[:31]

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    bold_font = Font(bold=True, size=12, color="000000")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    currency_font = Font(bold=True, size=13)
    custom_header_font = Font(bold=True, size=16, color="FFFFFF")
    custom_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_font = Font(bold=True, size=11, color="008000")
    grand_total_font = Font(bold=True, size=12, color="0000FF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Group operations by currency
    currencies = defaultdict(list)
    for op in data:
        currency_name = op.get("currency", {}).get("name", "Unknown")
        currencies[currency_name].append(op)

    row = 1
    currency_list = []
    total_in_base_currency = 0

    # Add custom header if provided
    if header_text:
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = header_text
        cell.font = custom_header_font
        cell.fill = custom_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 2

    # Add title
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = translator.get('total', lang).upper()
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    row += 2

    for currency_name, operations in currencies.items():
        # Sort operations by date
        operations.sort(key=lambda x: x.get("date", 0))

        # Add currency header
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = f"💱 {currency_name}"
        cell.font = currency_font
        cell.alignment = Alignment(horizontal='left')
        row += 1

        # Group by document_type
        doc_groups = defaultdict(list)
        for op in operations:
            doc_type_id = op.get("document_type", {}).get("id", "0")
            doc_groups[doc_type_id].append(op)

        total_currency_sum = 0

        for doc_type_id, ops in doc_groups.items():
            if not ops:
                continue

            # Calculate totals
            total_credit = sum(o.get("debit", 0) for o in ops)
            total_debit = sum(o.get("credit", 0) for o in ops)
            total_value = total_debit - total_credit
            total_currency_sum += total_value

            # Document type row
            doc_type_name = translator.get(f"plural_partner_document_type{doc_type_id}", lang)

            ws.cell(row=row, column=1).value = "📄"
            ws.cell(row=row, column=1).border = border

            ws.cell(row=row, column=2).value = doc_type_name
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')

            ws.cell(row=row, column=3).value = f"{total_value:,.2f}"
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')

            row += 1

        # Currency total row
        ws.cell(row=row, column=2).value = f"💰 {translator.get('currency_total', lang)}"
        ws.cell(row=row, column=2).font = total_font
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')

        ws.cell(row=row, column=3).value = f"{total_currency_sum:,.2f}"
        ws.cell(row=row, column=3).font = total_font
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=3).fill = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")

        row += 2

        # Calculate for base currency
        currency_id = operations[-1].get("currency", {}).get("id", "0")
        exchange_rate = operations[-1].get("currency", {}).get("exchange_rate", 1)
        start_amount = operations[0].get("start_amount", 0)

        currency_list.append({
            "id": currency_id,
            "name": currency_name,
            "exchange_rate": exchange_rate
        })

        total_in_base_currency += (total_currency_sum + start_amount) * exchange_rate

    # Add grand total section
    row += 1
    word_total = translator.get('total', lang)

    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = word_total.upper()
    cell.font = grand_total_font
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")
    row += 1

    for currency_info in currency_list:
        exchange_rate = currency_info['exchange_rate']

        ws.cell(row=row, column=2).value = f"{word_total} ({currency_info['name']})"
        ws.cell(row=row, column=2).font = bold_font
        ws.cell(row=row, column=2).border = border

        ws.cell(row=row, column=3).value = f"{total_in_base_currency / exchange_rate:,.2f}"
        ws.cell(row=row, column=3).font = bold_font
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"partner_balance_total_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Save workbook
    wb.save(filepath)

    return filepath

# Example usage:
# from core.conf import translator_service, TEST_INTEGRATION_TOKEN
# from regos.reports import RegosReports
# import asyncio
#
# regos_reports = RegosReports()
# data = asyncio.run(
#     regos_reports.partner_balance_report(
#         token=TEST_INTEGRATION_TOKEN,
#         partner_id=6,
#         firm_id=1,
#         start_time="01.01.2025 00:00:00",
#         end_time="12.11.2025 00:00:00"
#     )
# )
#
# # Export detailed balance
# filepath = format_partner_balance_excel(data=data["result"], lang="uz", translator=translator_service)
# print(f"Detailed report saved to: {filepath}")
#
# # Export totals
# filepath_total = format_total_excel(data=data["result"], lang="uz", translator=translator_service)
# print(f"Total report saved to: {filepath_total}")